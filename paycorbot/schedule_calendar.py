import json
import re
import sys
from datetime import datetime, timedelta
from bs4 import BeautifulSoup
import pandas as pd

def standardize_time_str(time_str):
    """Standardizes a time string to a uniform format for parsing."""
    time_str = time_str.strip().lower()
    time_str = time_str.replace('.', '').replace(' ', '')
    if time_str.endswith('a'):
        time_str = time_str[:-1] + 'am'
    elif time_str.endswith('p'):
        time_str = time_str[:-1] + 'pm'
    if not re.match(r'\d+:\d+', time_str):
        time_str = time_str[:-2] + ':00' + time_str[-2:]
    return time_str

def parse_time_str(time_str):
    """Parses a standardized time string into a datetime.time object."""
    time_str = standardize_time_str(time_str)
    time_formats = ['%I:%M%p', '%I%p']
    for fmt in time_formats:
        try:
            return datetime.strptime(time_str, fmt).time()
        except ValueError:
            continue
    raise ValueError(f"Time data '{time_str}' does not match any known format.")

class ScheduleDay:
    """Represents a single day's schedule."""
    def __init__(self, date, time=None, hours=0):
        self.date = date
        self.time = time
        self.hours = hours

    def __str__(self):
        return f"Date: {self.date}, Time: {self.time or 'OFF'}, Hours: {self.hours:.1f}"

    def to_dict(self):
        """Converts ScheduleDay to a dictionary for DataFrame construction."""
        if self.time:
            start_time_str, end_time_str = self.time.split("/")
            start_time = parse_time_str(start_time_str)
            end_time = parse_time_str(end_time_str)
            shift_start = datetime.combine(self.date, start_time)
            shift_end = datetime.combine(self.date, end_time)
            if shift_end <= shift_start:
                shift_end += timedelta(days=1)
            duration = round((shift_end - shift_start).total_seconds() / 3600, 2)
            return {
                "Shift Date": self.date,
                "Start Time": shift_start,
                "End Time": shift_end,
                "Duration (Hours)": duration,
            }
        else:
            return {
                "Shift Date": self.date,
                "Start Time": "OFF",
                "End Time": "OFF",
                "Duration (Hours)": 0,
            }

def parse_raw_markup(page_source):
    """Parses raw markup and extracts schedule information."""
    soup = BeautifulSoup(page_source, "html.parser")
    schedule_rows = soup.select("div.x-grid-item-container table")
    days = []

    for table in schedule_rows:
        rows = table.select("tr")
        for row in rows:
            cells = row.select("td")
            for cell in cells:
                date = cell.select_one(".indv-sch-cell-date-dom")
                time = cell.select_one(".indv-sch-sch-sten")
                hours = cell.select_one(".indv-sch-sch-hrs")
                if date and time and hours:
                    day_date = datetime(datetime.now().year, datetime.now().month, int(date.text.strip()))
                    day_time = time.text.strip()
                    day_hours = float(hours.text.strip().replace("h", ""))
                    days.append(ScheduleDay(day_date, day_time, day_hours))
    return days

def fill_days_off(shifts, start_date, end_date):
    """Fills in missing days as 'OFF' in the schedule."""
    all_dates = pd.date_range(start=start_date, end=end_date, freq="D").to_pydatetime().tolist()
    shift_dict = {shift.date: shift for shift in shifts}
    filled_shifts = []
    for date in all_dates:
        if date in shift_dict:
            filled_shifts.append(shift_dict[date])
        else:
            filled_shifts.append(ScheduleDay(date))
    return filled_shifts

def parse_shift_times(shift_times, date):
    """
    Parses the shift times and returns a dictionary with shift details.
    """
    try:
        start_time_str, end_time_str = shift_times.split('/')
        start_time = parse_time_str(start_time_str)
        end_time = parse_time_str(end_time_str)
        shift_start = datetime.combine(date.date(), start_time)
        shift_end = datetime.combine(date.date(), end_time)
        if shift_end <= shift_start:
            shift_end += timedelta(days=1)
        return {
            'Shift Date': date.date(),
            'Start Time': shift_start,
            'End Time': shift_end,
            'Duration (Hours)': round((shift_end - shift_start).total_seconds() / 3600, 2)
        }
    except ValueError as e:
        print(f"Error parsing times '{shift_times}': {e}")
        return None


from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

def format_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    header_font = Font(bold=True, size=12)
    for col in range(1, ws.max_column + 1):
        ws.cell(row=1, column=col).font = header_font

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center")

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

    formatted_file_path = file_path.replace(".xlsx", "_formatted.xlsx")
    wb.save(formatted_file_path)
    print(f"Formatted file saved as: {formatted_file_path}")

def output_to_excel(shifts, output_file):
    """Outputs the shifts to an Excel file."""
    shift_data = [shift.to_dict() for shift in shifts]
    df = pd.DataFrame(shift_data)

    df['Start Time'] = pd.to_datetime(df['Start Time'], errors="coerce")
    df['End Time'] = pd.to_datetime(df['End Time'], errors="coerce")

    df.to_excel(output_file, index=False)
    format_excel(output_file)
    print(f"Shifts have been successfully written to {output_file}")

def parse_calendar():
    """
    Main function to parse a schedule JSON file and output the data to an Excel file.
    This function uses argparse to handle command-line arguments for specifying the input JSON file
    and the output Excel file. It reads the JSON data from the input file, parses the schedule, and
    writes the parsed data to the output Excel file.
    Command-line arguments:
    --in : str : Path to the input JSON file (required).
    --out : str : Path to the output Excel file (required).
    Raises:
    FileNotFoundError: If the input JSON file does not exist.
    json.JSONDecodeError: If the input file is not a valid JSON.
    """
    parser = argparse.ArgumentParser(description='Parse schedule JSON and output to Excel.')
    parser.add_argument('--in', dest='input_file', required=True, help='Input JSON file')
    parser.add_argument('--out', dest='output_file', required=True, help='Output Excel file')
    args = parser.parse_args()
    with open(args.input_file, 'r', encoding='utf-8') as f:
        json_data = f.read()
    # Parser? I barely know her!
    shifts = parse_schedule(json_data)


    if shifts:
        output_to_excel(shifts, args.output_file)
    else:
        print('No shifts found in the schedule.')

if __name__ == '__main__':
    parse_calendar()
